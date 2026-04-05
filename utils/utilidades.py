import sys as _sys, os as _os
_ROOT = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
_sys.path.insert(0, _ROOT)
"""
utils/utilidades.py
Utilidades: validaciones, imágenes, exportación Excel/PDF
"""
import os
import re
import shutil
from datetime import datetime
from tkinter import filedialog, messagebox

# ✅ DEFINE BASE_DIR AQUÍ
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── Carpeta de exportaciones (dentro del proyecto) ──────────────
CARPETA_EXPORTS = os.path.join(BASE_DIR, "exports")
os.makedirs(CARPETA_EXPORTS, exist_ok=True)

# ── Pillow ──────────────────────────────────────────────────────
try:
    from PIL import Image, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ── Carpeta de imágenes del sistema ─────────────────────────────
CARPETA_IMG = os.path.join(BASE_DIR, "assets", "imagenes")
os.makedirs(CARPETA_IMG, exist_ok=True)

FORMATOS_OK = (".jpg", ".jpeg", ".png", ".gif")


# ══════════════════════════════════════════════════════════════════
# VALIDACIONES
# ══════════════════════════════════════════════════════════════════

def es_vacio(valor: str) -> bool:
    return valor.strip() == ""

def es_numero(valor: str) -> bool:
    try:
        float(valor.strip())
        return True
    except ValueError:
        return False

def es_entero(valor: str) -> bool:
    try:
        int(valor.strip())
        return True
    except ValueError:
        return False

def es_email_valido(valor: str) -> bool:
    return re.match(r'^[\w\.\+\-]+@[\w\-]+\.[a-zA-Z]{2,}$', valor.strip()) is not None

def validar_longitud(valor: str, minimo: int = 2, maximo: int = 100) -> bool:
    return minimo <= len(valor.strip()) <= maximo

def solo_letras_espacios(valor: str) -> bool:
    return bool(re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]+$', valor.strip()))

def bloquear_no_numericos(evento):
    caracter = evento.char
    if caracter and caracter not in '0123456789.,\x08\x7f\t':
        return "break"


# ══════════════════════════════════════════════════════════════════
# IMÁGENES  (Pillow)
# ══════════════════════════════════════════════════════════════════

def seleccionar_imagen(lbl_preview, var_ruta):
    if not PIL_OK:
        messagebox.showerror("Error", "Pillow no está instalado.\npip install Pillow")
        return
    ruta = filedialog.askopenfilename(
        title="Seleccionar imagen",
        filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif"), ("Todos", "*.*")]
    )
    if not ruta:
        return
    ext = os.path.splitext(ruta)[1].lower()
    if ext not in FORMATOS_OK:
        messagebox.showerror("Formato inválido",
                             f"Formato '{ext}' no permitido.\nUsa: JPG, PNG o GIF.")
        return
    tam_mb = os.path.getsize(ruta) / (1024 * 1024)
    if tam_mb > 5:
        messagebox.showerror("Imagen muy grande",
                             f"El archivo pesa {tam_mb:.1f} MB.\nMáximo permitido: 5 MB.")
        return
    var_ruta.set(ruta)
    try:
        img = Image.open(ruta)
        img.thumbnail((120, 100))
        foto = ImageTk.PhotoImage(img)
        lbl_preview.config(image=foto, text="")
        lbl_preview.image = foto
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la imagen:\n{e}")


def guardar_imagen(ruta_origen: str, nombre_archivo: str) -> str:
    if not PIL_OK or not ruta_origen:
        return ""
    try:
        ext = os.path.splitext(ruta_origen)[1].lower()
        if ext == ".gif":
            ext = ".png"
        destino = os.path.join(CARPETA_IMG, f"{nombre_archivo}{ext}")
        img = Image.open(ruta_origen)
        img.thumbnail((800, 600))
        img.save(destino)
        return destino
    except Exception as e:
        messagebox.showerror("Error al guardar imagen", str(e))
        return ""


def cargar_preview(lbl_preview, ruta: str):
    if not PIL_OK or not ruta or not os.path.exists(ruta):
        lbl_preview.config(image="", text="Sin imagen")
        return
    try:
        img = Image.open(ruta)
        img.thumbnail((120, 100))
        foto = ImageTk.PhotoImage(img)
        lbl_preview.config(image=foto, text="")
        lbl_preview.image = foto
    except Exception:
        lbl_preview.config(image="", text="Error imagen")


# ══════════════════════════════════════════════════════════════════
# HELPERS INTERNOS
# ══════════════════════════════════════════════════════════════════

def _nombre_archivo(titulo: str, ext: str) -> str:
    """Genera un nombre de archivo con timestamp para evitar sobreescrituras."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre = titulo.replace(" ", "_").lower()
    return os.path.join(CARPETA_EXPORTS, f"{nombre}_{ts}.{ext}")


# ══════════════════════════════════════════════════════════════════
# EXPORTAR A EXCEL  (openpyxl)
# ══════════════════════════════════════════════════════════════════

def exportar_excel(titulo: str, columnas: list, filas: list,
                   filtro_info: str = ""):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messagebox.showerror("Error", "Instala openpyxl:\npip install openpyxl")
        return

    ruta = _nombre_archivo(titulo, "xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    fill_header = PatternFill("solid", fgColor="1565C0")
    font_header = Font(bold=True, color="FFFFFF", size=11)
    font_titulo  = Font(bold=True, size=14)
    alineacion   = Alignment(horizontal="center", vertical="center")
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Título del reporte
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(columnas))
    celda_titulo = ws.cell(row=1, column=1, value=f"Reporte: {titulo}")
    celda_titulo.font = font_titulo
    celda_titulo.alignment = alineacion

    # Info de filtros
    if filtro_info:
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2,   end_column=len(columnas))
        ws.cell(row=2, column=1,
                value=f"Filtros: {filtro_info}").alignment = alineacion

    fila_header = 3 if filtro_info else 2

    # Encabezados
    for col_num, nombre in enumerate(columnas, 1):
        c = ws.cell(row=fila_header, column=col_num, value=nombre)
        c.font = font_header
        c.fill = fill_header
        c.alignment = alineacion
        c.border = borde

    # Datos con filas alternas
    fill_par   = PatternFill("solid", fgColor="E3F2FD")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")
    for idx, fila in enumerate(filas, 1):
        fill = fill_par if idx % 2 == 0 else fill_impar
        for col_num, valor in enumerate(fila, 1):
            c = ws.cell(row=fila_header + idx, column=col_num,
                        value=valor if valor is not None else "")
            c.fill = fill
            c.border = borde
            c.alignment = Alignment(horizontal="center")

    # Ajustar anchos de columna
    for col_num in range(1, len(columnas) + 1):
        letra = get_column_letter(col_num)
        max_ancho = max(
            len(str(ws.cell(row=r, column=col_num).value or ""))
            for r in range(fila_header, fila_header + len(filas) + 1)
        )
        ws.column_dimensions[letra].width = min(max(max_ancho + 4, 12), 40)

    wb.save(ruta)
    messagebox.showinfo(
        "✅ Excel exportado",
        f"Archivo guardado en:\n{ruta}"
    )


# ══════════════════════════════════════════════════════════════════
# EXPORTAR A PDF  (reportlab)
# ══════════════════════════════════════════════════════════════════

def exportar_pdf(titulo: str, columnas: list, filas: list,
                 filtro_info: str = ""):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        messagebox.showerror("Error", "Instala reportlab:\npip install reportlab")
        return

    ruta = _nombre_archivo(titulo, "pdf")

    doc = SimpleDocTemplate(
        ruta,
        pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm,     bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "Titulo", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1565C0"),
        spaceAfter=6, alignment=TA_CENTER
    )
    estilo_sub = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.grey,
        spaceAfter=12, alignment=TA_CENTER
    )

    elementos = []
    elementos.append(Paragraph(f"Reporte: {titulo}", estilo_titulo))
    if filtro_info:
        elementos.append(Paragraph(f"Filtros aplicados: {filtro_info}", estilo_sub))
    elementos.append(Spacer(1, 0.3*cm))

    datos = [columnas] + [
        [str(v)[:25] if v is not None else "" for v in fila]
        for fila in filas
    ]

    ancho_pagina = landscape(A4)[0] - 3*cm
    ancho_col = ancho_pagina / len(columnas)
    t = Table(datos, colWidths=[ancho_col] * len(columnas), repeatRows=1)

    estilo_tabla = TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor("#1565C0")),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, 0),  9),
        ("ALIGN",          (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#E3F2FD"), colors.white]),
        ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",       (0, 1), (-1, -1), 8),
        ("ALIGN",          (0, 1), (-1, -1), "CENTER"),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.HexColor("#90CAF9")),
        ("BOX",            (0, 0), (-1, -1), 1,   colors.HexColor("#1565C0")),
        ("ROWHEIGHT",      (0, 0), (-1, -1), 18),
    ])
    t.setStyle(estilo_tabla)
    elementos.append(t)

    doc.build(elementos)
    messagebox.showinfo(
        "✅ PDF exportado",
        f"Archivo guardado en:\n{ruta}"
    )