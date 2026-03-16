"""
Sistema ConstruData — Versión Completa
========================================
Dependencias:
    pip install openpyxl reportlab fpdf2 tkcalendar Pillow mysql-connector-python

Estructura de módulos:
    - Proyectos   (con imagen de portada)
    - Empleados   (con foto de perfil)
    - Materiales
    - Proveedores
    - Exportación (Excel + PDF con filtros)
"""

import re
import os
import shutil
import mysql.connector
from datetime import datetime
from tkinter import *
from tkinter import ttk, messagebox, filedialog

# ── Librerías opcionales con mensajes claros si faltan ──────────────────────
try:
    from tkcalendar import DateEntry
    TKCALENDAR_OK = True
except ImportError:
    TKCALENDAR_OK = False
    print("⚠ tkcalendar no instalado. Ejecuta: pip install tkcalendar")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("⚠ openpyxl no instalado. Ejecuta: pip install openpyxl")

try:
    from fpdf import FPDF
    FPDF_OK = True
except ImportError:
    FPDF_OK = False
    print("⚠ fpdf2 no instalado. Ejecuta: pip install fpdf2")

try:
    from PIL import Image, ImageTk, ImageFilter
    PIL_OK = True
except ImportError:
    PIL_OK = False
    print("⚠ Pillow no instalado. Ejecuta: pip install Pillow")

# ── Carpeta para imágenes del sistema ───────────────────────────────────────
IMG_DIR = os.path.join(os.path.dirname(__file__), "construdata_imagenes")
os.makedirs(IMG_DIR, exist_ok=True)

# ════════════════════════════════════════════════════════════════════════════
#  CONEXIÓN
# ════════════════════════════════════════════════════════════════════════════

def conectar():
    return mysql.connector.connect(
        host="localhost", user="root", password="", database="construdata"
    )

# ════════════════════════════════════════════════════════════════════════════
#  VALIDACIONES
# ════════════════════════════════════════════════════════════════════════════

def validar_numero(valor, campo="Campo", entero=False):
    """Devuelve (True, valor_convertido) o (False, mensaje_error)."""
    v = valor.strip()
    if not v:
        return False, f"{campo} es obligatorio."
    try:
        return (True, int(v)) if entero else (True, float(v))
    except ValueError:
        tipo = "entero" if entero else "numérico"
        return False, f"{campo} debe ser un valor {tipo}. Recibido: '{v}'"


def validar_texto(valor, campo="Campo", min_len=2, max_len=100):
    v = valor.strip()
    if len(v) < min_len:
        return False, f"{campo} debe tener al menos {min_len} caracteres."
    if len(v) > max_len:
        return False, f"{campo} no puede superar {max_len} caracteres."
    if re.search(r'[<>"\';]', v):
        return False, f"{campo} contiene caracteres no permitidos (< > \" ' ;)."
    return True, v


def validar_email(valor):
    patron = r'^[\w\.\+\-]+@[\w\-]+\.[a-zA-Z]{2,}$'
    if not re.match(patron, valor.strip()):
        return False, f"Email inválido: '{valor}'. Formato esperado: usuario@dominio.com"
    return True, valor.strip()


def validar_fecha(valor, campo="Fecha"):
    """Acepta YYYY-MM-DD o DD/MM/YYYY."""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return True, datetime.strptime(valor.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return False, f"{campo} inválida. Use YYYY-MM-DD o DD/MM/YYYY."


def mostrar_error(msg):
    messagebox.showerror("Error de validación", msg)


def confirmar(pregunta):
    return messagebox.askyesno("Confirmar operación", pregunta)

# ════════════════════════════════════════════════════════════════════════════
#  MANEJO DE IMÁGENES (Pillow)
# ════════════════════════════════════════════════════════════════════════════

FORMATOS_PERMITIDOS = {".jpg", ".jpeg", ".png", ".gif"}
TAMANO_MAX_MB = 5
THUMB_SIZE = (120, 120)


def validar_imagen(ruta):
    """Verifica formato y tamaño. Devuelve (True, ruta) o (False, mensaje)."""
    if not PIL_OK:
        return False, "Pillow no está instalado."
    ext = os.path.splitext(ruta)[1].lower()
    if ext not in FORMATOS_PERMITIDOS:
        return False, f"Formato no permitido: '{ext}'. Use JPG, PNG o GIF."
    tam_mb = os.path.getsize(ruta) / (1024 * 1024)
    if tam_mb > TAMANO_MAX_MB:
        return False, f"La imagen pesa {tam_mb:.1f} MB. Máximo permitido: {TAMANO_MAX_MB} MB."
    return True, ruta


def guardar_imagen(ruta_origen, nombre_destino):
    """
    Copia la imagen al directorio de imágenes del sistema,
    la redimensiona a 800x600 máximo y devuelve la ruta destino.
    """
    ext = os.path.splitext(ruta_origen)[1].lower()
    nombre_archivo = f"{nombre_destino}{ext}"
    ruta_destino = os.path.join(IMG_DIR, nombre_archivo)

    img = Image.open(ruta_origen)
    # Convertir GIF animado a PNG estático
    if img.format == "GIF":
        img = img.convert("RGBA")
        nombre_archivo = f"{nombre_destino}.png"
        ruta_destino = os.path.join(IMG_DIR, nombre_archivo)
    # Redimensionar manteniendo proporción
    img.thumbnail((800, 600), Image.LANCZOS)
    img.save(ruta_destino)
    return ruta_destino


def crear_thumbnail(ruta, size=THUMB_SIZE):
    """Devuelve un PhotoImage para mostrar en la UI."""
    if not PIL_OK:
        return None
    try:
        img = Image.open(ruta)
        if img.format == "GIF":
            img = img.convert("RGBA")
        img.thumbnail(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        return None


def selector_imagen(label_preview, var_ruta):
    """Abre diálogo, valida y muestra preview de imagen."""
    ruta = filedialog.askopenfilename(
        title="Seleccionar imagen",
        filetypes=[("Imágenes", "*.jpg *.jpeg *.png *.gif"), ("Todos", "*.*")]
    )
    if not ruta:
        return
    ok, msg = validar_imagen(ruta)
    if not ok:
        mostrar_error(msg)
        return
    var_ruta.set(ruta)
    thumb = crear_thumbnail(ruta)
    if thumb:
        label_preview.config(image=thumb, text="")
        label_preview.image = thumb   # evitar garbage collection

# ════════════════════════════════════════════════════════════════════════════
#  WIDGETS REUTILIZABLES
# ════════════════════════════════════════════════════════════════════════════

def campo_texto(parent, etiqueta, fila, col=0, ancho=22, min_len=2, max_len=100):
    Label(parent, text=etiqueta).grid(row=fila, column=col, sticky="w", padx=8, pady=3)
    e = Entry(parent, width=ancho)
    e.grid(row=fila, column=col + 1, padx=4, pady=3, sticky="w")
    return e


def campo_numero(parent, etiqueta, fila, col=0, ancho=14):
    Label(parent, text=etiqueta).grid(row=fila, column=col, sticky="w", padx=8, pady=3)
    e = Entry(parent, width=ancho)
    e.grid(row=fila, column=col + 1, padx=4, pady=3, sticky="w")
    # Bloquear caracteres no numéricos en tiempo real
    def solo_numeros(event):
        val = e.get()
        limpio = re.sub(r'[^0-9\.\-]', '', val)
        if val != limpio:
            pos = e.index(INSERT)
            e.delete(0, END)
            e.insert(0, limpio)
            e.icursor(min(pos, len(limpio)))
    e.bind("<KeyRelease>", solo_numeros)
    return e


def campo_fecha(parent, etiqueta, fila, col=0):
    Label(parent, text=etiqueta).grid(row=fila, column=col, sticky="w", padx=8, pady=3)
    if TKCALENDAR_OK:
        e = DateEntry(parent, width=14, date_pattern="yyyy-mm-dd",
                      background="#2d6a4f", foreground="white", borderwidth=2)
        e.grid(row=fila, column=col + 1, padx=4, pady=3, sticky="w")
    else:
        e = Entry(parent, width=14)
        e.insert(0, "YYYY-MM-DD")
        e.grid(row=fila, column=col + 1, padx=4, pady=3, sticky="w")
    return e


def obtener_fecha(widget):
    """Obtiene la fecha de un DateEntry o Entry normal."""
    if TKCALENDAR_OK and isinstance(widget, DateEntry):
        return widget.get_date().strftime("%Y-%m-%d")
    return widget.get()


def tabla_con_scroll(parent, columnas, alto=12):
    frame = Frame(parent)
    frame.pack(fill="both", expand=True, padx=8, pady=6)

    tv = ttk.Treeview(frame, columns=columnas, show="headings", height=alto)
    for col in columnas:
        tv.heading(col, text=col)
        tv.column(col, width=max(80, len(col) * 9), anchor="center")

    sb_v = ttk.Scrollbar(frame, orient="vertical",   command=tv.yview)
    sb_h = ttk.Scrollbar(frame, orient="horizontal", command=tv.xview)
    tv.configure(yscrollcommand=sb_v.set, xscrollcommand=sb_h.set)

    sb_v.pack(side="right",  fill="y")
    sb_h.pack(side="bottom", fill="x")
    tv.pack(side="left", fill="both", expand=True)
    return tv


def limpiar_tabla(tv):
    for i in tv.get_children():
        tv.delete(i)


def cargar_tabla(tv, filas):
    limpiar_tabla(tv)
    for fila in filas:
        tv.insert("", END, values=[v if v is not None else "" for v in fila])

# ════════════════════════════════════════════════════════════════════════════
#  EXPORTACIÓN — EXCEL
# ════════════════════════════════════════════════════════════════════════════

ESTILO_HEADER = {
    "font":      Font(bold=True, color="FFFFFF", size=11),
    "fill":      PatternFill("solid", fgColor="1F4E79"),
    "alignment": Alignment(horizontal="center", vertical="center"),
    "border":    Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
}


def exportar_excel(titulo, columnas, filas, nombre_archivo="reporte.xlsx"):
    if not OPENPYXL_OK:
        mostrar_error("openpyxl no instalado. Ejecuta: pip install openpyxl")
        return

    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=nombre_archivo,
        title="Guardar Excel"
    )
    if not ruta:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    # Título del reporte
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(columnas))
    celda_titulo = ws.cell(row=1, column=1, value=f"CONSTRUDATA — {titulo.upper()}")
    celda_titulo.font      = Font(bold=True, size=14, color="FFFFFF")
    celda_titulo.fill      = PatternFill("solid", fgColor="0D3B66")
    celda_titulo.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Fecha generación
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=len(columnas))
    ws.cell(row=2, column=1,
            value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    # Encabezados
    for ci, col in enumerate(columnas, 1):
        c = ws.cell(row=3, column=ci, value=col)
        c.font      = ESTILO_HEADER["font"]
        c.fill      = ESTILO_HEADER["fill"]
        c.alignment = ESTILO_HEADER["alignment"]
        c.border    = ESTILO_HEADER["border"]
    ws.row_dimensions[3].height = 20

    # Datos con filas alternas
    fill_par  = PatternFill("solid", fgColor="D6E4F0")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")
    borde = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )
    for ri, fila in enumerate(filas, 4):
        relleno = fill_par if ri % 2 == 0 else fill_impar
        for ci, valor in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=valor)
            c.fill   = relleno
            c.border = borde
            c.alignment = Alignment(horizontal="left")

    # Autoajustar columnas
    for col_idx, col_name in enumerate(columnas, 1):
        max_w = len(str(col_name))
        for fila in filas:
            v = str(fila[col_idx - 1]) if col_idx - 1 < len(fila) else ""
            max_w = max(max_w, len(v))
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = min(max_w + 4, 40)

    # Total de registros
    fila_total = len(filas) + 4
    ws.cell(row=fila_total, column=1,
            value=f"Total de registros: {len(filas)}")
    ws.cell(row=fila_total, column=1).font = Font(bold=True)

    wb.save(ruta)
    messagebox.showinfo("Excel exportado", f"Archivo guardado en:\n{ruta}")


# ════════════════════════════════════════════════════════════════════════════
#  EXPORTACIÓN — PDF
# ════════════════════════════════════════════════════════════════════════════

def exportar_pdf(titulo, columnas, filas, nombre_archivo="reporte.pdf"):
    if not FPDF_OK:
        mostrar_error("fpdf2 no instalado. Ejecuta: pip install fpdf2")
        return

    ruta = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF", "*.pdf")],
        initialfile=nombre_archivo,
        title="Guardar PDF"
    )
    if not ruta:
        return

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Encabezado
    pdf.set_fill_color(13, 59, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, f"CONSTRUDATA — {titulo.upper()}", ln=True, align="C", fill=True)

    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
             f"Total registros: {len(filas)}", ln=True, align="R")
    pdf.ln(3)

    # Calcular ancho de columnas
    page_w = pdf.w - 2 * pdf.l_margin
    col_w  = page_w / len(columnas)

    # Cabeceras de tabla
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    for col in columnas:
        pdf.cell(col_w, 8, str(col)[:20], border=1, align="C", fill=True)
    pdf.ln()

    # Filas de datos con colores alternos
    pdf.set_font("Helvetica", "", 8)
    for ri, fila in enumerate(filas):
        if ri % 2 == 0:
            pdf.set_fill_color(214, 228, 240)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 30, 30)
        for valor in fila:
            texto = str(valor)[:25] if valor is not None else ""
            pdf.cell(col_w, 7, texto, border=1, align="L", fill=True)
        pdf.ln()

    # Pie de página
    pdf.set_y(-15)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, f"Sistema ConstruData — Página {pdf.page_no()}", align="C")

    pdf.output(ruta)
    messagebox.showinfo("PDF exportado", f"Archivo guardado en:\n{ruta}")

# ════════════════════════════════════════════════════════════════════════════
#  MÓDULO PROYECTOS
# ════════════════════════════════════════════════════════════════════════════

def build_proyectos(parent):
    # ── Panel de formulario ──────────────────────────────────────────────
    paned = PanedWindow(parent, orient=HORIZONTAL)
    paned.pack(fill="both", expand=True)

    frm_form = LabelFrame(paned, text="Datos del Proyecto", padx=8, pady=8)
    paned.add(frm_form, minsize=360)

    frm_right = Frame(paned)
    paned.add(frm_right)

    # Campos del formulario
    e_cod   = campo_texto(frm_form,  "Código *",          0, max_len=20)
    e_nom   = campo_texto(frm_form,  "Nombre *",          1, ancho=28, max_len=100)
    e_dir   = campo_texto(frm_form,  "Dirección",         2, ancho=28)
    e_tipo  = campo_texto(frm_form,  "Tipo construcción", 3, ancho=22)
    e_area  = campo_numero(frm_form, "Área total (m²)",   4)
    e_pisos = campo_numero(frm_form, "Pisos",             5)
    e_ini   = campo_fecha(frm_form,  "Fecha inicio",      6)
    e_fin   = campo_fecha(frm_form,  "Fecha fin est.",    7)
    e_est   = campo_texto(frm_form,  "Estado",            8, max_len=50)
    e_pre   = campo_numero(frm_form, "Presupuesto",       9)
    e_ger   = campo_texto(frm_form,  "Gerente",           10, ancho=28)

    # ── Imagen de portada del proyecto ──────────────────────────────────
    frm_img = LabelFrame(frm_form, text="Imagen de portada", padx=6, pady=6)
    frm_img.grid(row=11, column=0, columnspan=4, sticky="ew", pady=6, padx=4)

    lbl_preview = Label(frm_img, text="Sin imagen",
                        width=16, height=7, relief="groove", bg="#f0f0f0")
    lbl_preview.pack(side=LEFT, padx=6)

    var_img_ruta = StringVar()

    frm_img_btns = Frame(frm_img)
    frm_img_btns.pack(side=LEFT, padx=6)
    Button(frm_img_btns, text="📁 Seleccionar imagen",
           command=lambda: selector_imagen(lbl_preview, var_img_ruta)
           ).pack(anchor="w", pady=2)
    Button(frm_img_btns, text="🗑 Quitar imagen",
           command=lambda: [var_img_ruta.set(""),
                            lbl_preview.config(image="", text="Sin imagen")]
           ).pack(anchor="w", pady=2)
    Label(frm_img_btns, text="Formatos: JPG, PNG, GIF\nMáx. 5 MB",
          fg="gray", font=("Arial", 8)).pack(anchor="w")

    # ── Filtros de exportación ───────────────────────────────────────────
    frm_filtros = LabelFrame(frm_form, text="Filtros de exportación", padx=6, pady=4)
    frm_filtros.grid(row=12, column=0, columnspan=4, sticky="ew", pady=4, padx=4)

    Label(frm_filtros, text="Estado:").grid(row=0, column=0, sticky="w")
    cmb_estado = ttk.Combobox(frm_filtros, width=14,
                               values=["Todos", "En proceso", "Planeacion", "Finalizado"])
    cmb_estado.set("Todos")
    cmb_estado.grid(row=0, column=1, padx=4)

    Label(frm_filtros, text="Desde:").grid(row=0, column=2, padx=(8,2))
    f_desde = campo_fecha(frm_filtros, "", 0, col=2)
    f_hasta = campo_fecha(frm_filtros, "Hasta:", 0, col=4)

    # ── Tabla ────────────────────────────────────────────────────────────
    cols = ("ID","Código","Nombre","Dirección","Tipo","Área","Pisos",
            "Inicio","Fin Est.","Estado","Presupuesto","Gerente")
    tv = tabla_con_scroll(frm_right, cols, alto=14)

    # ── Barra de botones ─────────────────────────────────────────────────
    frm_btns = Frame(frm_right)
    frm_btns.pack(fill="x", padx=8, pady=4)

    def limpiar():
        for e in [e_cod, e_nom, e_dir, e_tipo, e_area, e_pisos, e_est, e_pre, e_ger]:
            e.delete(0, END)
        var_img_ruta.set("")
        lbl_preview.config(image="", text="Sin imagen")

    def registrar():
        # Validaciones
        ok, cod  = validar_texto(e_cod.get(),  "Código",  1, 20)
        if not ok: return mostrar_error(cod)
        ok, nom  = validar_texto(e_nom.get(),  "Nombre",  2, 100)
        if not ok: return mostrar_error(nom)

        area_val = e_area.get().strip()
        if area_val:
            ok, area_val = validar_numero(area_val, "Área")
            if not ok: return mostrar_error(area_val)

        pisos_val = e_pisos.get().strip()
        if pisos_val:
            ok, pisos_val = validar_numero(pisos_val, "Pisos", entero=True)
            if not ok: return mostrar_error(pisos_val)

        pre_val = e_pre.get().strip()
        if pre_val:
            ok, pre_val = validar_numero(pre_val, "Presupuesto")
            if not ok: return mostrar_error(pre_val)

        fecha_ini = obtener_fecha(e_ini)
        fecha_fin = obtener_fecha(e_fin)

        try:
            con = conectar(); cur = con.cursor()
            cur.execute("SELECT id_proyecto FROM proyectos WHERE codigo=%s", (cod,))
            if cur.fetchone():
                con.close()
                return mostrar_error(f"Ya existe un proyecto con código '{cod}'.")

            # Guardar imagen si hay
            ruta_guardada = None
            if var_img_ruta.get():
                ruta_guardada = guardar_imagen(var_img_ruta.get(), f"proyecto_{cod}")

            cur.execute("""
                INSERT INTO proyectos
                    (codigo, nombre, direccion, tipo_construccion,
                     area_total, cantidad_pisos, fecha_inicio,
                     fecha_fin_estimada, estado, presupuesto, gerente_proyecto)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (cod, nom, e_dir.get(), e_tipo.get(),
                  area_val or None, pisos_val or None,
                  fecha_ini, fecha_fin,
                  e_est.get(), pre_val or None, e_ger.get()))
            con.commit(); con.close()
            messagebox.showinfo("Éxito", "Proyecto registrado correctamente.")
            limpiar(); ver()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def eliminar():
        sel = tv.selection()
        if not sel:
            return mostrar_error("Selecciona un proyecto de la tabla.")
        valores = tv.item(sel[0])["values"]
        if not confirmar(f"¿Eliminar el proyecto '{valores[2]}'?\nEsta acción no se puede deshacer."):
            return
        try:
            con = conectar(); cur = con.cursor()
            cur.execute("DELETE FROM proyectos WHERE id_proyecto=%s", (valores[0],))
            con.commit(); con.close()
            messagebox.showinfo("Eliminado", "Proyecto eliminado.")
            ver()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def ver():
        try:
            con = conectar(); cur = con.cursor()
            estado_f = cmb_estado.get()
            if estado_f == "Todos":
                cur.execute("SELECT * FROM proyectos")
            else:
                cur.execute("SELECT * FROM proyectos WHERE estado=%s", (estado_f,))
            cargar_tabla(tv, cur.fetchall())
            con.close()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def get_filas_exportar():
        try:
            con = conectar(); cur = con.cursor()
            estado_f = cmb_estado.get()
            if estado_f == "Todos":
                cur.execute("SELECT * FROM proyectos")
            else:
                cur.execute("SELECT * FROM proyectos WHERE estado=%s", (estado_f,))
            return cur.fetchall()
        except Exception:
            return []

    for txt, cmd, color in [
        ("💾 Registrar",  registrar,  "#1a6e3c"),
        ("🔍 Ver / Filtrar", ver,    "#0d5aa7"),
        ("🗑 Eliminar",   eliminar,   "#b52b2b"),
        ("🧹 Limpiar",    limpiar,    "#555555"),
        ("📊 Excel",      lambda: exportar_excel(
            "Proyectos", list(cols), get_filas_exportar(), "proyectos.xlsx"), "#217346"),
        ("📄 PDF",        lambda: exportar_pdf(
            "Proyectos", list(cols), get_filas_exportar(), "proyectos.pdf"),  "#c0392b"),
    ]:
        Button(frm_btns, text=txt, command=cmd, bg=color, fg="white",
               width=14, relief="flat", cursor="hand2").pack(side=LEFT, padx=3)

    ver()

# ════════════════════════════════════════════════════════════════════════════
#  MÓDULO EMPLEADOS
# ════════════════════════════════════════════════════════════════════════════

def build_empleados(parent):
    paned = PanedWindow(parent, orient=HORIZONTAL)
    paned.pack(fill="both", expand=True)

    frm_form = LabelFrame(paned, text="Datos del Empleado", padx=8, pady=8)
    paned.add(frm_form, minsize=340)
    frm_right = Frame(paned)
    paned.add(frm_right)

    e_dni  = campo_texto(frm_form,  "DNI *",            0, max_len=20)
    e_nom  = campo_texto(frm_form,  "Nombres *",        1, ancho=24)
    e_ape  = campo_texto(frm_form,  "Apellidos *",      2, ancho=24)
    e_nac  = campo_fecha(frm_form,  "Fecha nacimiento", 3)
    e_dir  = campo_texto(frm_form,  "Dirección",        4, ancho=28)
    e_tel  = campo_texto(frm_form,  "Teléfono",         5, max_len=20)
    e_mail = campo_texto(frm_form,  "Email *",          6, ancho=28)
    e_exp  = campo_numero(frm_form, "Años experiencia", 7)
    e_sal  = campo_numero(frm_form, "Salario",          8)
    e_cont = campo_texto(frm_form,  "Tipo contrato",    9, max_len=50)
    e_car  = campo_texto(frm_form,  "Cargo",            10, max_len=50)
    e_fcon = campo_fecha(frm_form,  "Fecha contratación",11)

    # ── Foto de perfil ───────────────────────────────────────────────────
    frm_img = LabelFrame(frm_form, text="Foto de perfil", padx=6, pady=6)
    frm_img.grid(row=12, column=0, columnspan=4, sticky="ew", pady=6, padx=4)

    lbl_foto = Label(frm_img, text="Sin foto",
                     width=14, height=7, relief="groove", bg="#f0f0f0")
    lbl_foto.pack(side=LEFT, padx=6)

    var_foto = StringVar()
    frm_fb = Frame(frm_img)
    frm_fb.pack(side=LEFT, padx=6)
    Button(frm_fb, text="📁 Seleccionar foto",
           command=lambda: selector_imagen(lbl_foto, var_foto)).pack(anchor="w", pady=2)
    Button(frm_fb, text="🗑 Quitar",
           command=lambda: [var_foto.set(""),
                            lbl_foto.config(image="", text="Sin foto")]
           ).pack(anchor="w", pady=2)
    Label(frm_fb, text="JPG, PNG, GIF — Máx. 5 MB",
          fg="gray", font=("Arial", 8)).pack(anchor="w")

    # ── Filtros ──────────────────────────────────────────────────────────
    frm_fil = LabelFrame(frm_form, text="Filtro exportación", padx=6, pady=4)
    frm_fil.grid(row=13, column=0, columnspan=4, sticky="ew", pady=4, padx=4)
    Label(frm_fil, text="Cargo:").grid(row=0, column=0, sticky="w")
    e_fil_cargo = Entry(frm_fil, width=18)
    e_fil_cargo.grid(row=0, column=1, padx=4)
    Label(frm_fil, text="(vacío = todos)", fg="gray", font=("Arial", 8)).grid(row=0, column=2, sticky="w")

    # ── Tabla ────────────────────────────────────────────────────────────
    cols = ("ID","DNI","Nombres","Apellidos","Nacimiento","Dirección",
            "Teléfono","Email","Exp.","Salario","Contrato","Cargo","F.Contratación")
    tv = tabla_con_scroll(frm_right, cols, alto=14)

    frm_btns = Frame(frm_right)
    frm_btns.pack(fill="x", padx=8, pady=4)

    def limpiar():
        for e in [e_dni, e_nom, e_ape, e_dir, e_tel, e_mail,
                  e_exp, e_sal, e_cont, e_car]:
            e.delete(0, END)
        var_foto.set("")
        lbl_foto.config(image="", text="Sin foto")

    def registrar():
        ok, dni_v = validar_texto(e_dni.get(), "DNI", 5, 20)
        if not ok: return mostrar_error(dni_v)
        ok, nom_v = validar_texto(e_nom.get(), "Nombres")
        if not ok: return mostrar_error(nom_v)
        ok, ape_v = validar_texto(e_ape.get(), "Apellidos")
        if not ok: return mostrar_error(ape_v)
        ok, mail_v = validar_email(e_mail.get())
        if not ok: return mostrar_error(mail_v)

        exp_v = e_exp.get().strip()
        if exp_v:
            ok, exp_v = validar_numero(exp_v, "Años experiencia", entero=True)
            if not ok: return mostrar_error(exp_v)

        sal_v = e_sal.get().strip()
        if sal_v:
            ok, sal_v = validar_numero(sal_v, "Salario")
            if not ok: return mostrar_error(sal_v)

        try:
            con = conectar(); cur = con.cursor()
            cur.execute("SELECT id_empleado FROM empleados WHERE dni=%s", (dni_v,))
            if cur.fetchone():
                con.close()
                return mostrar_error(f"Ya existe un empleado con DNI '{dni_v}'.")

            if var_foto.get():
                guardar_imagen(var_foto.get(), f"empleado_{dni_v}")

            cur.execute("""
                INSERT INTO empleados
                    (dni, nombres, apellidos, fecha_nacimiento, direccion,
                     telefono, anios_experiencia, salario, tipo_contrato,
                     cargo, fecha_contratacion)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (dni_v, nom_v, ape_v,
                  obtener_fecha(e_nac), e_dir.get(), e_tel.get(),
                  exp_v or None, sal_v or None,
                  e_cont.get(), e_car.get(), obtener_fecha(e_fcon)))
            con.commit(); con.close()
            messagebox.showinfo("Éxito", "Empleado registrado.")
            limpiar(); ver()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def eliminar():
        sel = tv.selection()
        if not sel:
            return mostrar_error("Selecciona un empleado de la tabla.")
        v = tv.item(sel[0])["values"]
        if not confirmar(f"¿Eliminar al empleado '{v[2]} {v[3]}'?"):
            return
        try:
            con = conectar(); cur = con.cursor()
            cur.execute("DELETE FROM empleados WHERE id_empleado=%s", (v[0],))
            con.commit(); con.close()
            messagebox.showinfo("Eliminado", "Empleado eliminado.")
            ver()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def ver():
        try:
            con = conectar(); cur = con.cursor()
            cargo_f = e_fil_cargo.get().strip()
            if cargo_f:
                cur.execute("""SELECT id_empleado,dni,nombres,apellidos,
                                      fecha_nacimiento,direccion,telefono,
                                      anios_experiencia,salario,tipo_contrato,
                                      cargo,fecha_contratacion
                               FROM empleados WHERE cargo LIKE %s""",
                            (f"%{cargo_f}%",))
            else:
                cur.execute("""SELECT id_empleado,dni,nombres,apellidos,
                                      fecha_nacimiento,direccion,telefono,
                                      anios_experiencia,salario,tipo_contrato,
                                      cargo,fecha_contratacion
                               FROM empleados""")
            cargar_tabla(tv, cur.fetchall())
            con.close()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def get_filas():
        con = conectar(); cur = con.cursor()
        cargo_f = e_fil_cargo.get().strip()
        if cargo_f:
            cur.execute("SELECT * FROM empleados WHERE cargo LIKE %s",
                        (f"%{cargo_f}%",))
        else:
            cur.execute("SELECT * FROM empleados")
        filas = cur.fetchall(); con.close()
        return filas

    for txt, cmd, color in [
        ("💾 Registrar",     registrar, "#1a6e3c"),
        ("🔍 Ver / Filtrar", ver,       "#0d5aa7"),
        ("🗑 Eliminar",      eliminar,  "#b52b2b"),
        ("🧹 Limpiar",       limpiar,   "#555555"),
        ("📊 Excel",         lambda: exportar_excel(
            "Empleados", list(cols), get_filas(), "empleados.xlsx"), "#217346"),
        ("📄 PDF",           lambda: exportar_pdf(
            "Empleados", list(cols), get_filas(), "empleados.pdf"),  "#c0392b"),
    ]:
        Button(frm_btns, text=txt, command=cmd, bg=color, fg="white",
               width=14, relief="flat", cursor="hand2").pack(side=LEFT, padx=3)

    ver()

# ════════════════════════════════════════════════════════════════════════════
#  MÓDULO MATERIALES
# ════════════════════════════════════════════════════════════════════════════

def build_materiales(parent):
    paned = PanedWindow(parent, orient=HORIZONTAL)
    paned.pack(fill="both", expand=True)

    frm_form = LabelFrame(paned, text="Datos del Material", padx=8, pady=8)
    paned.add(frm_form, minsize=300)
    frm_right = Frame(paned)
    paned.add(frm_right)

    e_cod  = campo_texto(frm_form,  "Código *",         0, max_len=20)
    e_des  = campo_texto(frm_form,  "Descripción *",    1, ancho=28)
    e_cat  = campo_texto(frm_form,  "Categoría",        2, max_len=50)
    e_uni  = campo_texto(frm_form,  "Unidad medida",    3, max_len=20)
    e_esp  = campo_texto(frm_form,  "Especificaciones", 4, ancho=28, max_len=200)
    e_pro  = campo_numero(frm_form, "ID Proveedor",     5)

    # Filtros exportación
    frm_fil = LabelFrame(frm_form, text="Filtro exportación", padx=6, pady=4)
    frm_fil.grid(row=6, column=0, columnspan=4, sticky="ew", pady=4, padx=4)
    Label(frm_fil, text="Categoría:").grid(row=0, column=0, sticky="w")
    e_fil_cat = Entry(frm_fil, width=18)
    e_fil_cat.grid(row=0, column=1, padx=4)

    cols = ("ID","Código","Descripción","Categoría","Unidad","Especificaciones","ID Proveedor")
    tv = tabla_con_scroll(frm_right, cols, alto=14)
    frm_btns = Frame(frm_right)
    frm_btns.pack(fill="x", padx=8, pady=4)

    def limpiar():
        for e in [e_cod, e_des, e_cat, e_uni, e_esp, e_pro]:
            e.delete(0, END)

    def registrar():
        ok, cod = validar_texto(e_cod.get(), "Código", 1, 20)
        if not ok: return mostrar_error(cod)
        ok, des = validar_texto(e_des.get(), "Descripción")
        if not ok: return mostrar_error(des)

        pro_v = e_pro.get().strip()
        if pro_v:
            ok, pro_v = validar_numero(pro_v, "ID Proveedor", entero=True)
            if not ok: return mostrar_error(pro_v)

        try:
            con = conectar(); cur = con.cursor()
            cur.execute("SELECT id_material FROM materiales WHERE codigo=%s", (cod,))
            if cur.fetchone():
                con.close()
                return mostrar_error(f"Ya existe un material con código '{cod}'.")
            cur.execute("""
                INSERT INTO materiales
                    (codigo, descripcion, categoria, unidad_medida,
                     especificaciones, proveedor_preferido)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (cod, des, e_cat.get(), e_uni.get(),
                  e_esp.get(), pro_v or None))
            con.commit(); con.close()
            messagebox.showinfo("Éxito", "Material registrado.")
            limpiar(); ver()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def eliminar():
        sel = tv.selection()
        if not sel: return mostrar_error("Selecciona un material.")
        v = tv.item(sel[0])["values"]
        if not confirmar(f"¿Eliminar el material '{v[2]}'?"): return
        try:
            con = conectar(); cur = con.cursor()
            cur.execute("DELETE FROM materiales WHERE id_material=%s", (v[0],))
            con.commit(); con.close()
            messagebox.showinfo("Eliminado", "Material eliminado.")
            ver()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def ver():
        try:
            con = conectar(); cur = con.cursor()
            cat_f = e_fil_cat.get().strip()
            if cat_f:
                cur.execute("SELECT * FROM materiales WHERE categoria LIKE %s",
                            (f"%{cat_f}%",))
            else:
                cur.execute("SELECT * FROM materiales")
            cargar_tabla(tv, cur.fetchall())
            con.close()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def get_filas():
        con = conectar(); cur = con.cursor()
        cat_f = e_fil_cat.get().strip()
        if cat_f:
            cur.execute("SELECT * FROM materiales WHERE categoria LIKE %s",
                        (f"%{cat_f}%",))
        else:
            cur.execute("SELECT * FROM materiales")
        filas = cur.fetchall(); con.close(); return filas

    for txt, cmd, color in [
        ("💾 Registrar",     registrar, "#1a6e3c"),
        ("🔍 Ver / Filtrar", ver,       "#0d5aa7"),
        ("🗑 Eliminar",      eliminar,  "#b52b2b"),
        ("🧹 Limpiar",       limpiar,   "#555555"),
        ("📊 Excel",         lambda: exportar_excel(
            "Materiales", list(cols), get_filas(), "materiales.xlsx"), "#217346"),
        ("📄 PDF",           lambda: exportar_pdf(
            "Materiales", list(cols), get_filas(), "materiales.pdf"),  "#c0392b"),
    ]:
        Button(frm_btns, text=txt, command=cmd, bg=color, fg="white",
               width=14, relief="flat", cursor="hand2").pack(side=LEFT, padx=3)

    ver()

# ════════════════════════════════════════════════════════════════════════════
#  MÓDULO PROVEEDORES
# ════════════════════════════════════════════════════════════════════════════

def build_proveedores(parent):
    paned = PanedWindow(parent, orient=HORIZONTAL)
    paned.pack(fill="both", expand=True)

    frm_form = LabelFrame(paned, text="Datos del Proveedor", padx=8, pady=8)
    paned.add(frm_form, minsize=280)
    frm_right = Frame(paned)
    paned.add(frm_right)

    e_nom  = campo_texto(frm_form, "Nombre *",    0, ancho=26, max_len=100)
    e_tel  = campo_texto(frm_form, "Teléfono",    1, max_len=20)
    e_dir  = campo_texto(frm_form, "Dirección",   2, ancho=26)
    e_mail = campo_texto(frm_form, "Email *",     3, ancho=26)

    cols = ("ID","Nombre","Teléfono","Dirección","Email")
    tv = tabla_con_scroll(frm_right, cols, alto=16)
    frm_btns = Frame(frm_right)
    frm_btns.pack(fill="x", padx=8, pady=4)

    def limpiar():
        for e in [e_nom, e_tel, e_dir, e_mail]:
            e.delete(0, END)

    def registrar():
        ok, nom_v = validar_texto(e_nom.get(), "Nombre")
        if not ok: return mostrar_error(nom_v)
        ok, mail_v = validar_email(e_mail.get())
        if not ok: return mostrar_error(mail_v)

        try:
            con = conectar(); cur = con.cursor()
            cur.execute("SELECT id_proveedor FROM proveedores WHERE nombre=%s AND telefono=%s",
                        (nom_v, e_tel.get()))
            if cur.fetchone():
                con.close()
                return mostrar_error("Ya existe ese proveedor.")
            cur.execute("""
                INSERT INTO proveedores (nombre, telefono, direccion, email)
                VALUES (%s,%s,%s,%s)
            """, (nom_v, e_tel.get(), e_dir.get(), mail_v))
            con.commit(); con.close()
            messagebox.showinfo("Éxito", "Proveedor registrado.")
            limpiar(); ver()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def eliminar():
        sel = tv.selection()
        if not sel: return mostrar_error("Selecciona un proveedor.")
        v = tv.item(sel[0])["values"]
        if not confirmar(f"¿Eliminar al proveedor '{v[1]}'?"): return
        try:
            con = conectar(); cur = con.cursor()
            cur.execute("DELETE FROM proveedores WHERE id_proveedor=%s", (v[0],))
            con.commit(); con.close()
            messagebox.showinfo("Eliminado", "Proveedor eliminado.")
            ver()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def ver():
        try:
            con = conectar(); cur = con.cursor()
            cur.execute("SELECT id_proveedor, nombre, telefono, direccion, email FROM proveedores")
            cargar_tabla(tv, cur.fetchall())
            con.close()
        except mysql.connector.Error as err:
            mostrar_error(str(err))

    def get_filas():
        con = conectar(); cur = con.cursor()
        cur.execute("SELECT id_proveedor, nombre, telefono, direccion, email FROM proveedores")
        filas = cur.fetchall(); con.close(); return filas

    for txt, cmd, color in [
        ("💾 Registrar",  registrar, "#1a6e3c"),
        ("🔍 Ver todos",  ver,       "#0d5aa7"),
        ("🗑 Eliminar",   eliminar,  "#b52b2b"),
        ("🧹 Limpiar",    limpiar,   "#555555"),
        ("📊 Excel",      lambda: exportar_excel(
            "Proveedores", list(cols), get_filas(), "proveedores.xlsx"), "#217346"),
        ("📄 PDF",        lambda: exportar_pdf(
            "Proveedores", list(cols), get_filas(), "proveedores.pdf"),  "#c0392b"),
    ]:
        Button(frm_btns, text=txt, command=cmd, bg=color, fg="white",
               width=14, relief="flat", cursor="hand2").pack(side=LEFT, padx=3)

    ver()

# ════════════════════════════════════════════════════════════════════════════
#  VENTANA PRINCIPAL
# ════════════════════════════════════════════════════════════════════════════

ventana = Tk()
ventana.title("Sistema ConstruData")
ventana.geometry("1200x700")
ventana.resizable(True, True)

# Barra de título
hdr = Frame(ventana, bg="#0d3b66", height=44)
hdr.pack(fill="x")
Label(hdr, text="  🏗  SISTEMA CONSTRUDATA — Gestión de Construcción",
      bg="#0d3b66", fg="white",
      font=("Segoe UI", 13, "bold"), pady=8).pack(side=LEFT)
Label(hdr, text=f"  {datetime.now().strftime('%d/%m/%Y')}  ",
      bg="#0d3b66", fg="#aac4e0",
      font=("Segoe UI", 9)).pack(side=RIGHT)

# Estado de librerías
libs = []
if not TKCALENDAR_OK: libs.append("tkcalendar")
if not OPENPYXL_OK:   libs.append("openpyxl")
if not FPDF_OK:       libs.append("fpdf2")
if not PIL_OK:        libs.append("Pillow")
if libs:
    aviso = Frame(ventana, bg="#fff3cd")
    aviso.pack(fill="x")
    Label(aviso,
          text=f"⚠  Librerías faltantes: {', '.join(libs)}  |  "
               f"Instala con: pip install {' '.join(libs)}",
          bg="#fff3cd", fg="#856404",
          font=("Segoe UI", 9), pady=3).pack()

# Notebook
style = ttk.Style()
style.configure("TNotebook.Tab", font=("Segoe UI", 10, "bold"), padding=[12, 6])

nb = ttk.Notebook(ventana)
nb.pack(fill="both", expand=True)

for titulo, builder in [
    ("🏗  Proyectos",   build_proyectos),
    ("👷  Empleados",   build_empleados),
    ("🧱  Materiales",  build_materiales),
    ("🚚  Proveedores", build_proveedores),
]:
    f = Frame(nb)
    nb.add(f, text=titulo)
    builder(f)

# Barra de estado
Frame(ventana, height=1, bg="#cccccc").pack(fill="x")
Label(ventana,
      text="  Conectado a: construdata@localhost  |  "
           f"Imágenes guardadas en: {IMG_DIR}",
      anchor="w", fg="#555555", font=("Segoe UI", 8), pady=3
      ).pack(fill="x")

ventana.mainloop()